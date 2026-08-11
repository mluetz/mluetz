#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VDA-ISA-Katalog-Konverter — ENX-Excel -> data/isa_catalog.json -> isa/index.html

Liest den offiziellen VDA-ISA-Prüfkatalog (ENX-Excel, getestet mit ISA6 EN 6.0.3)
aus data/, extrahiert Module, Kapitel und Controls in eine strukturierte JSON
und baut daraus die eigenständige Assessment-App isa/index.html
(Vorlage: tools/isa_template.html, Platzhalter /*__DATA__*/).

Katalog-Update (z. B. 6.0.4): neue Excel nach data/ legen, ISA_XLSX unten
anpassen (oder Pfad als Argument übergeben) und das Skript erneut ausführen.

  pip install openpyxl
  python3 tools/isa_convert.py [pfad/zur/ISA.xlsx]
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
ISA_XLSX = ROOT / "data" / "ISA6_EN_6.0.3.xlsx"
OUT_JSON = ROOT / "data" / "isa_catalog.json"
TEMPLATE = ROOT / "tools" / "isa_template.html"
OUT_HTML = ROOT / "isa" / "index.html"

CONTROL_RE = re.compile(r"^\d+\.\d+\.\d+$")
FILLER_RE = re.compile(r"intentional\s+invisible\s+text", re.I)

# Welche Blätter sind Katalog-Module? (Blattname, Modul-Kürzel, deutscher Anzeigename)
MODULES = [
    ("Information Security", "IS", "Informationssicherheit"),
    ("Prototype Protection", "PP", "Prototypenschutz"),
    ("Data Protection", "DP", "Datenschutz"),
]

# Spaltenerkennung über Header-Text (EN-Katalog; DE-Muster vorsorglich ergänzt).
# Reihenfolge ist relevant: erster Treffer gewinnt, "very high" vor "high" prüfen.
HEADER_PATTERNS = [
    ("question", r"control question|kontrollfrage"),
    ("objective", r"^objective|^ziel"),
    ("must", r"requirements.*\(must\)|anforderungen.*\(muss\)"),
    ("should", r"\(should\)|\(sollte\)"),
    ("veryHigh", r"very high protection|sehr hohen? schutzbedarf"),
    ("high", r"high protection needs?$|hohen? schutzbedarf$|high protection needs\b(?!.*very)"),
    ("vehicles", r"vehicles classified|fahrzeuge.*klassifiziert"),
    ("sga", r"simplified group assessments|vereinfachte.*gruppen"),
    ("responsible", r"person responsible|prozessverantwortlich"),
    ("refStd", r"reference to other standards|referenz.*standards"),
    ("refGuide", r"implementation guidance|umsetzungshinweise"),
    ("info", r"further information|weitere informationen"),
    ("exNormal", r"normal protection"),
    ("exHigh", r"examples.*high protection(?!.*very)|hoher schutzbedarf"),
    ("exVeryHigh", r"examples.*very high|sehr hoher schutzbedarf"),
    ("questions", r"possible questions|mögliche fragen"),
    ("evidence", r"possible evidence|mögliche nachweise"),
]


def norm_header(v):
    return re.sub(r"\s+", " ", str(v or "")).strip().lower()


def clean_text(v):
    """Zellwert säubern; Füll- und 'None'-Texte entfernen."""
    if v is None:
        return ""
    s = str(v).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not s or s.lower() == "none" or FILLER_RE.search(s):
        return ""
    return s


def parse_items(text):
    """Anforderungs-Zelle in Aufzählung parsen.

    '+' beginnt eine Hauptanforderung, '-' einen Unterpunkt, andere Zeilen
    werden an den letzten Punkt angehängt (Zeilenumbruch innerhalb eines Punkts).
    Rückgabe: Liste von {t: Text, s: [Untertexte]}.
    """
    text = clean_text(text)
    if not text:
        return []
    items = []
    for raw in text.split("\n"):
        line = raw.strip()
        if not line:
            continue
        if line.startswith("+"):
            items.append({"t": line[1:].strip(), "s": []})
        elif re.match(r"^[-–•]\s", line) or line in ("-", "–"):
            body = re.sub(r"^[-–•]\s*", "", line)
            if items:
                (items[-1]["s"]).append(body)
            else:
                items.append({"t": body, "s": []})
        else:
            if items:
                if items[-1]["s"]:
                    items[-1]["s"][-1] += " " + line
                else:
                    items[-1]["t"] += " " + line
            else:
                items.append({"t": line, "s": []})
    return items


def find_header_row(ws):
    """Zeile mit 'Row_Format' in Spalte A/B suchen (IS/PP: Zeile 2, DP: Zeile 4)."""
    for row in ws.iter_rows(min_row=1, max_row=8):
        for c in row[:3]:
            if norm_header(c.value) == "row_format":
                return c.row
    raise SystemExit(f"Header-Zeile (Row_Format) in Blatt '{ws.title}' nicht gefunden")


def map_columns(ws, header_row):
    cols = {}
    for c in ws[header_row]:
        h = norm_header(c.value)
        if not h:
            continue
        for key, pat in HEADER_PATTERNS:
            if key not in cols and re.search(pat, h):
                cols[key] = c.column
                break
    for required in ("question", "must"):
        if required not in cols:
            raise SystemExit(f"Spalte '{required}' in Blatt '{ws.title}' nicht gefunden")
    return cols


def cell(row_map, cols, key):
    col = cols.get(key)
    return clean_text(row_map.get(col)) if col else ""


def parse_sheet(ws, module_id):
    header_row = find_header_row(ws)
    cols = map_columns(ws, header_row)
    # Spalten A..C: Row_Format / Is_Title? / Control number (fix im ENX-Layout)
    chapters = []
    current_chapter = None
    current_sub = None
    controls = []

    for row in ws.iter_rows(min_row=header_row + 1):
        row_map = {i + 1: c.value for i, c in enumerate(row)}
        row_format = clean_text(row_map.get(1))
        cno = clean_text(row_map.get(3))
        title = cell(row_map, cols, "question")
        if row_format == "header":
            if not cno:
                continue
            depth = cno.count(".")
            if depth == 0:
                current_chapter = {"no": cno, "title": title, "subs": []}
                chapters.append(current_chapter)
                current_sub = None
            else:
                current_sub = {"no": cno, "title": title, "controls": []}
                if current_chapter:
                    current_chapter["subs"].append(current_sub)
        elif CONTROL_RE.match(cno):
            ctl = {
                "id": cno,
                "module": module_id,
                "chapter": current_chapter["no"] if current_chapter else cno.split(".")[0],
                "sub": current_sub["no"] if current_sub else None,
                "question": title,
                "objective": cell(row_map, cols, "objective"),
                "req": {},
                "responsible": cell(row_map, cols, "responsible"),
                "refStd": cell(row_map, cols, "refStd"),
                "refGuide": cell(row_map, cols, "refGuide"),
                "info": cell(row_map, cols, "info"),
                "examples": {
                    "normal": cell(row_map, cols, "exNormal"),
                    "high": cell(row_map, cols, "exHigh"),
                    "veryHigh": cell(row_map, cols, "exVeryHigh"),
                },
                "questions": cell(row_map, cols, "questions"),
                "evidence": cell(row_map, cols, "evidence"),
            }
            for req_key in ("must", "should", "high", "veryHigh", "vehicles", "sga"):
                col = cols.get(req_key)
                if col:
                    items = parse_items(row_map.get(col))
                    if items:
                        ctl["req"][req_key] = items
            if current_sub is not None:
                current_sub["controls"].append(cno)
            elif current_chapter is not None:
                # Kapitel ohne Unterkapitel-Header: implizites Unterkapitel anlegen
                current_sub = {"no": cno.rsplit(".", 1)[0], "title": "", "controls": [cno]}
                current_chapter["subs"].append(current_sub)
            controls.append(ctl)
    return chapters, controls


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else ISA_XLSX
    if not xlsx.exists():
        raise SystemExit(f"Katalog-Excel nicht gefunden: {xlsx}")
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    version = ""
    if "Cover" in wb.sheetnames:
        for row in wb["Cover"].iter_rows(min_row=1, max_row=40):
            for c in row:
                m = re.search(r"version:\s*([\d.]+)", str(c.value or ""), re.I)
                if m:
                    version = m.group(1)
    catalog = {
        "source": xlsx.name,
        "version": version or "6.0.3",
        "language": "en",
        "attribution": (
            "VDA ISA (Information Security Assessment), Version {v} — "
            "© ENX Association / Verband der Automobilindustrie e. V. (VDA), "
            "lizenziert unter CC BY-ND 4.0."
        ).format(v=version or "6.0.3"),
        "modules": [],
        "controls": {},
        "maturity": [],
    }
    if "Maturity levels" in wb.sheetnames:
        ws = wb["Maturity levels"]
        rows = {clean_text(r[1].value).lower(): r for r in ws.iter_rows(min_row=1, max_row=12) if len(r) > 1}
        names = rows.get("name")
        principles = rows.get("principle")
        definitions = rows.get("definition")
        for lvl in range(6):
            col = 2 + lvl  # C..H (0-basiert in der Zeilen-Tupel: Index 2..7)
            catalog["maturity"].append({
                "level": lvl,
                "name": clean_text(names[col].value) if names else "",
                "principle": clean_text(principles[col].value) if principles else "",
                "definition": clean_text(definitions[col].value) if definitions else "",
            })
    for sheet, mid, name_de in MODULES:
        if sheet not in wb.sheetnames:
            print(f"Hinweis: Blatt '{sheet}' fehlt — Modul {mid} übersprungen")
            continue
        chapters, controls = parse_sheet(wb[sheet], mid)
        catalog["modules"].append(
            {"id": mid, "sheet": sheet, "name": name_de, "chapters": chapters}
        )
        for ctl in controls:
            catalog["controls"][ctl["id"]] = ctl
        print(f"{sheet}: {len(controls)} Controls, {len(chapters)} Kapitel")

    OUT_JSON.write_text(
        json.dumps(catalog, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    print(f"-> {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size // 1024} KB)")

    if TEMPLATE.exists():
        html = TEMPLATE.read_text(encoding="utf-8")
        payload = json.dumps(catalog, ensure_ascii=False, separators=(",", ":"))
        # </script> im JSON-Payload für das Einbetten in HTML entschärfen
        payload = payload.replace("</", "<\\/")
        if "/*__DATA__*/" not in html:
            raise SystemExit("Platzhalter /*__DATA__*/ fehlt in der Vorlage")
        html = html.replace("/*__DATA__*/", payload)
        OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
        OUT_HTML.write_text(html, encoding="utf-8")
        print(f"-> {OUT_HTML.relative_to(ROOT)} ({OUT_HTML.stat().st_size // 1024} KB)")
    else:
        print(f"Vorlage {TEMPLATE.relative_to(ROOT)} fehlt — nur JSON erzeugt")


if __name__ == "__main__":
    main()
