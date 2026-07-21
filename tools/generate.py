#!/usr/bin/env python3
"""
Generator für die TISAX-AL3-Finding-Datenbank (REUTIB / Reutter-Group).

Liest das Excel-Finding-Register, extrahiert das Blatt "TISAX Finding Register",
leitet Domänen-Tags (u. a. OT) und Cross-Referenzen ab und erzeugt aus
tools/template.html die auslieferbare, selbsttragende index.html.

Aufruf:
    python3 tools/generate.py [pfad/zur/quelle.xlsx]

Ohne Argument wird data/TISAX_Finding_Register_v3_M_N_O.xlsx verwendet.
Erzeugt:  index.html  und  data/findings.json
Abhängigkeit: openpyxl  (pip install openpyxl)
"""
import json
import re
import sys
import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SRC = ROOT / "data" / "TISAX_Finding_Register_v3_M_N_O.xlsx"
TEMPLATE = ROOT / "tools" / "template.html"
OUT_HTML = ROOT / "index.html"
OUT_JSON = ROOT / "data" / "findings.json"
SHEET = "TISAX Finding Register"

# Spaltenreihenfolge im Register (A..W)
KEYS = ["findingId", "isaModul", "kapitel", "kontrollNr", "frageEn", "frageDe",
        "typ", "beschreibung", "reifeIst", "reifeSoll", "gap", "massnahme",
        "umsetzung", "xref", "scope", "verantwortlich", "standort", "prioritaet",
        "status", "frist", "nachweis", "isbKommentar", "ct"]

MODNAME = {
    "1": "1 · IS Policies & Organisation",
    "2": "2 · Human Resources",
    "3": "3 · Physical Security",
    "4": "4 · Identity & Access Mgmt",
    "5": "5 · IT(OT) Security / Cybersecurity",
    "6": "6 · Supplier Relationships",
    "7": "7 · Compliance",
}


def conv(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if v is None:
        return ""
    return v.strip() if isinstance(v, str) else v


def module_of(kn):
    m = re.match(r"\s*(\d+)\.", kn or "")
    return m.group(1) if m else None


def derive_tags(rec):
    """Leitet Domänen-Tags (Themenfilter) aus Modul und Inhalt ab."""
    blob = " ".join(str(rec.get(k, "")) for k in
                    ["kapitel", "verantwortlich", "umsetzung", "massnahme",
                     "beschreibung", "frageDe", "scope"]).lower()
    mod = module_of(rec["kontrollNr"])
    t = []
    ot = ("it(ot)" in (rec["kapitel"] or "").lower()
          or "ot/prod" in blob
          or re.search(r"\bot\b", (rec["verantwortlich"] or "").lower())
          or "62443" in blob or mod == "5")
    if ot:
        t.append("OT")
    if any(w in blob for w in ["bcm", "notfall", "verfügbarkeit", "continuity",
                               "dr-test", " krisen", "ausfall", "redundan", "rto", "rpo"]):
        t.append("BCM/Verfügbarkeit")
    if any(w in blob for w in ["datenschutz", "dsgvo", " dsb", "personenbezogen"]):
        t.append("Datenschutz")
    if any(w in blob for w in ["supplier", "lieferant", "3rd party", "third party", "dienstleister"]):
        t.append("Supplier")
    if mod == "2":
        t.append("HR")
    if mod == "3":
        t.append("Physical")
    if mod == "4":
        t.append("IAM")
    if mod == "1":
        t.append("Governance/Policy")
    if any(w in blob for w in ["malware", "virus", "antivir"]):
        t.append("Malware")
    if any(w in blob for w in ["logging", "siem", " event", "protokollier", "monitoring"]):
        t.append("Logging/Monitoring")
    if mod == "7":
        t.append("Compliance")
    return sorted(set(t))


def parse_xrefs(rec):
    """Findet 'Row NN / F-XXX'-Verweise in den Spalten N und M.

    Kategorisiert nach den Abschnitts-Überschriften der Spalte N:
    'pre' = vorgelagert (Voraussetzung), 'post' = nachgelagert (liefert Input an),
    'cross' = Querverweis bzw. unklassifiziert.
    """
    out, seen = [], set()

    def scan(txt, kind):
        for line in (txt or "").split("\n"):
            low = line.lower()
            if "vorgelagert" in low:
                kind = "pre"
            elif "nachgelagert" in low:
                kind = "post"
            elif "querverweis" in low:
                kind = "cross"
            for m in re.finditer(r"Row\s+(\d+)\s*/\s*(F-[\w]+)", line):
                key = (int(m.group(1)), m.group(2))
                if key not in seen:
                    seen.add(key)
                    out.append({"row": int(m.group(1)), "fid": m.group(2), "kind": kind})

    scan(rec["xref"], "cross")
    scan(rec["umsetzung"], "cross")
    return out


def extract(src: Path):
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb[SHEET]
    data = []
    for r in range(2, ws.max_row + 1):
        vals = [conv(ws.cell(r, c).value) for c in range(1, 24)]
        if all((v == "" or v is None) for v in vals):
            continue
        rec = dict(zip(KEYS, vals))
        rec["excelRow"] = r
        rec["module"] = module_of(rec["kontrollNr"]) or ""
        rec["moduleName"] = MODNAME.get(rec["module"], "General / Katalog")
        rec["tags"] = derive_tags(rec)
        rec["xrefs"] = parse_xrefs(rec)
        data.append(rec)
    return data


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
    data = extract(src)
    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")

    tpl = TEMPLATE.read_text(encoding="utf-8")
    if "/*__DATA__*/" not in tpl:
        raise SystemExit("Template-Platzhalter /*__DATA__*/ fehlt.")
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    OUT_HTML.write_text(tpl.replace("/*__DATA__*/", payload), encoding="utf-8")

    ot = sum(1 for d in data if "OT" in d["tags"])
    print(f"OK · {len(data)} Findings · {ot} OT-relevant · {OUT_HTML.name} ({OUT_HTML.stat().st_size//1024} KB)")


if __name__ == "__main__":
    main()
